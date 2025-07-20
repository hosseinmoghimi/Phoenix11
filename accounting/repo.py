from .models import Category,FinancialDocument,FinancialDocumentLine,InvoiceLineItemUnit
from .models import InvoiceLine,InvoiceLineItem,Account,Product,Service,FinancialEvent,FinancialYear
from .models import Invoice,Bank,BankAccount,PersonCategory,FinancialYear,PersonAccount,ProductSpecification
from .models import Brand
from .apps import APP_NAME
from .enums import *
from log.repo import LogRepo
from django.db.models import Q
from django.shortcuts import reverse
from authentication.repo import ProfileRepo
# from processmanagement.permission import Permission,OperationEnum
from utility.num import filter_number
from utility.calendar import PersianCalendar
from utility.constants import FAILED,SUCCEED
from utility.log import leolog
from .constants import EXCEL_PRODUCTS_DATA_START_ROW,EXCEL_SERVICES_DATA_START_ROW
from .defaults import default_accounts,default_persons,default_banks
from .enums import AccountTypeEnum,AccountNatureEnum
from .settings_on_server import ACCOUNT_LEVEL_NAMES
# from processmanagement.models import Permission
class InvoiceLineItemUnitRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        
        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_event"):
            self.objects=InvoiceLineItemUnit.objects
        elif request.user.is_authenticated:
            accs=[]
            for person in Person.objects.filter(profile__user_id=request.user.id):

                my_accounts=AccountRepo(request=request).my_accounts
                for acc in my_accounts:
                    accs.append(acc.id)
            self.objects=Event.objects.filter(Q(bedehkar_id__in=accs)|Q(bestankar_id__in=accs))
        else:
            self.objects=Event.objects.filter(pk=0)

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
            
        if "invoice_line_item_id" in kwargs:
            objects=objects.filter(invoice_line_item_id=kwargs['invoice_line_item_id']) 
        return objects.all()
    def product_unit(self,*args, **kwargs):
        if "product_unit_id" in kwargs:
            return self.objects.filter(pk=kwargs['product_unit_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

                     
    def add_invoice_line_item_unit(self,*args, **kwargs):
        invoice_line_item_unit,message,result=(None,"",FAILED)
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"productprice"):
        if not self.request.user.has_perm(APP_NAME+".add_invoicelineitemunit"):
            message="دسترسی غیر مجاز"
            return result,message,invoice_line_item_unit
        
 
        invoice_line_item_unit=InvoiceLineItemUnit()
        if 'invoice_line_item_id' in kwargs:
            invoice_line_item_unit.invoice_line_item_id=kwargs['invoice_line_item_id']

        if 'service' in kwargs:
            invoice_line_item_unit.invoice_line_item_id=kwargs['service']
        
        if 'product_id' in kwargs:
            invoice_line_item_unit.invoice_line_item_id=kwargs['product_id']
        if 'unit_price' in kwargs:
            invoice_line_item_unit.unit_price=kwargs['unit_price']
        if 'unit_name' in kwargs:
            invoice_line_item_unit.unit_name=kwargs['unit_name']
        if 'default' in kwargs:
            invoice_line_item_unit.default=kwargs['default']
        if 'coef' in kwargs:
            invoice_line_item_unit.coef=kwargs['coef']
        if invoice_line_item_unit.unit_price<1:
            message='قیمت را صفر انتخاب کرده اید.'
            return result,message,invoice_line_item_unit
        invoice_line_item_unit.save()
        result=SUCCEED
        message="قیمت جدید با موفقیت اضافه گردید."
         
        invoice_line_item_units=InvoiceLineItemUnit.objects.filter(invoice_line_item_id=invoice_line_item_unit.invoice_line_item.id)
        return result,message,invoice_line_item_units


class InvoiceLineRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        
        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_event"):
            self.objects=InvoiceLine.objects
        elif request.user.is_authenticated:
            accs=[]
            for person in Person.objects.filter(profile__user_id=request.user.id):

                my_accounts=AccountRepo(request=request).my_accounts
                for acc in my_accounts:
                    accs.append(acc.id)
            self.objects=InvoiceLine.objects.filter(Q(bedehkar_id__in=accs)|Q(bestankar_id__in=accs))
        else:
            self.objects=InvoiceLine.objects.filter(pk=0)

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
        return objects.all()
    def invoice_line(self,*args, **kwargs):
        if "invoice_line_id" in kwargs:
            return self.objects.filter(pk=kwargs['invoice_line_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

            
        
    def add_invoice_line(self,*args,**kwargs):
        result,message,meal=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_invoiceline"):
            message="دسترسی غیر مجاز"
            return result,message,meal

        invoice_line=InvoiceLine()
        if 'invoice_line_item_id' in kwargs:
            invoice_line_item_id=kwargs["invoice_line_item_id"]
            invoice_line.invoice_line_item_id=invoice_line_item_id
 
        if 'invoice_id' in kwargs:
            invoice_line.invoice_id=kwargs["invoice_id"]
        if 'discount_percentage' in kwargs:
            invoice_line.discount_percentage=kwargs["discount_percentage"]
        if 'quantity' in kwargs:
            invoice_line.quantity=kwargs["quantity"]
        if 'unit_price' in kwargs:
            unit_price=kwargs["unit_price"]
            invoice_line.unit_price=unit_price

        if 'unit_name' in kwargs:
            unit_name=kwargs["unit_name"]
            invoice_line.unit_name=unit_name
        if 'save' in kwargs:
            save=kwargs["save"]
            if save:
                if 'coef' in kwargs:
                    coef=kwargs["coef"]
                if 'default' in kwargs:
                    default11=kwargs["default"]
                InvoiceLineItemUnitRepo(request=self.request).add_invoice_line_item_unit(
                    invoice_line_item_id=invoice_line_item_id,
                    coef=coef,
                    default=default11,
                    unit_name=unit_name,
                    unit_price=unit_price,
                    )

        result=SUCCEED
        message="سطر فاکتور با موفقیت اضافه شد."     
 
        invoice_line.save()
        return result,message,invoice_line
           

class AccountRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Account.objects.filter(id=0)
        profile=ProfileRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_account"):
                self.objects=Account.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            codeee=str(filter_number(search_for))
            objects=objects.filter(Q(title__contains=search_for) | Q(code=search_for) | Q(code=codeee) )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "level" in kwargs:
            level=kwargs["level"]
            objects=objects.filter(level=level)  
        return objects.all()
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def account(self,*args, **kwargs):
        if "account_id" in kwargs and kwargs["account_id"] is not None:
            return self.objects.filter(pk=kwargs['account_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            code=kwargs['code']
            code=kwargs['code']

            account= self.objects.filter(code=code).first()
            return account
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
       

           
    def set_account_parent(self,*args, **kwargs):
        result,message,account,parent=FAILED,"",None,None
        account=self.account(*args,**kwargs)
        parent=None
        parent_code=kwargs["parent_code"]
        if parent_code=='0':
            account.parent=None
        else:    
            parent=self.account(account_code=parent_code)
            account.parent_id=parent.id
        account.save()
        result=SUCCEED
        message="با موفقیت تغییر یافت"
        return result,message,account,parent


    def initial_default_accounts(self,*args, **kwargs):
        account_group_counter=0
        basic_accounts_counter=0
        moein_accounts_counter=0
        moein2_accounts_counter=0
        tafsili_accounts_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_account"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        
        accounts=default_accounts() 

        for account in accounts:
            parent_account=None
            if 'parent_code' in account:
                parent_account=Account.objects.filter(code=account["parent_code"]).first()
            new_account=Account(title=account["name"],color=account["color"],code=account['code'],priority=account['priority'],parent=parent_account)
            # new_account=Account(parent=parent_account,**kwargs)
            new_account.save()
            # account_group_counter+=1
            # basic_accounts_counter+=1   
 
                                  
        accounts=Account.objects
        account_group_counter=len(accounts.filter(type=AccountTypeEnum.GROUP))
        basic_accounts_counter=len(accounts.filter(type=AccountTypeEnum.BASIC))
        moein_accounts_counter=len(accounts.filter(type=AccountTypeEnum.MOEIN_1))
        moein2_accounts_counter=len(accounts.filter(type=AccountTypeEnum.MOEIN_2))
        tafsili_accounts_counter=len(accounts.filter(Q(type=AccountTypeEnum.TAFSILI_1)|Q(type=AccountTypeEnum.TAFSILI_2)|Q(type=AccountTypeEnum.TAFSILI_3)|Q(type=AccountTypeEnum.TAFSILI_4)))
                                                        

        if result==SUCCEED:
            message="با موفقیت اضافه گردید."
        message+=f"<br>{account_group_counter}   گروه حساب" 
        message+=f"<br>{basic_accounts_counter}   حساب  کل " 
        message+=f"<br>{moein_accounts_counter}  حساب معین سطح یک " 
        message+=f"<br>{moein2_accounts_counter}  حساب معین سطح دو " 
        message+=f"<br>{tafsili_accounts_counter}  حساب تفصیلی " 

        me_profile=ProfileRepo(request=self.request).me
        new_log={}
        new_log['title']="افزودن حساب های پیش فرض"
        new_log['app_name']=APP_NAME
        new_log['profile']=me_profile
        new_log['description']="حساب های پیش فرض اضافه شدند."
        LogRepo(request=self.request).add_log(**new_log)
        return result,message
 
    def delete_all_accounts(self,*args, **kwargs):
        if not self.request.user.has_perm(APP_NAME+".delete_account"):
            message="دسترسی غیر مجاز"
            return message,result
        FinancialDocumentLine.objects.all().delete()
        FinancialEvent.objects.all().delete()
        FinancialDocument.objects.all().delete()
        # TafsiliAccount.objects.all().delete()
        # MoeinAccount.objects.all().delete()
        # BasicAccount.objects.all().delete()
        # AccountGroup.objects.all().delete() 
        Account.objects.all().delete() 
        result=SUCCEED
        message="همه حساب ها حذف شد."

        me_profile=ProfileRepo(request=self.request).me
        new_log={}
        new_log['title']="حذف همه حساب ها"
        new_log['app_name']=APP_NAME
        new_log['profile']=me_profile
        new_log['description']="همه ی حساب ها حذف شدند."
        LogRepo(request=self.request).add_log(**new_log)

        return result,message
    
    def add_account(self,*args,**kwargs):
        result,message,account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_account"):
            message="دسترسی غیر مجاز"
            return result,message,account

        account=Account()
        if 'title' in kwargs:
            account.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                account.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            account.color=kwargs["color"]
        if 'code' in kwargs:
            account.code=kwargs["code"]
        if 'priority' in kwargs:
            account.priority=kwargs["priority"]
        if 'type' in kwargs:
            account.type=kwargs["type"]

            
        if 'parent_code' in kwargs:
            parent_code= kwargs["parent_code"]
            parent=Account.objects.filter(code=parent_code).first()
            if parent is not None:
                account.parent_id=parent.id

        if 'nature' in kwargs:
            account.nature=kwargs["nature"]
        (result,message,account)=account.save()
        return result,message,account


class PersonAccountRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=PersonAccount.objects 
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(category__contains=search_for)   )

        if "category" in kwargs:
            category=kwargs["category"]
            objects=objects.filter(Q(category=category)   )
        return objects.all()
  
    def add_person_account(self,*args, **kwargs):
        result,message,person_account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message,person_account
        person_account=PersonAccount()
        if 'color' in kwargs:
            person_account.color=kwargs['color']
        if 'title' in kwargs:
            person_account.title=kwargs['title']
        if 'code' in kwargs:
            code=kwargs['code']
            person_account.code=code

        if 'person' in kwargs:
            person_account.person=kwargs['person']
        if 'person_id' in kwargs:
            person_account.person_id=kwargs['person_id']
        if 'person_category' in kwargs:
            person_account.person_category=kwargs['person_category']
        if 'person_category_id' in kwargs:
            person_account.person_category_id=kwargs['person_category_id']
        result,message,person_account=person_account.save()
        # result=SUCCEED
        # message="با موفقیت حساب فرد ایجاد شد."
        
        return result,message,person_account

    def person_account(self,*args, **kwargs):
        if "person_account_id" in kwargs and kwargs["person_account"] is not None:
            return self.objects.filter(pk=kwargs['person_account_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
             
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
                    
    def delete_all(self,*args,**kwargs):
        
        if not self.request.user.has_perm(APP_NAME+".delete_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message
        PersonAccount.objects.all().delete() 
                   
        result=SUCCEED
        message="همه حذف شدند."
        return result,message
     

    def add_account_to_person(self,*args,**kwargs):
        result,message,account,person,action=FAILED,"",None,None,None
        if not self.request.user.has_perm(APP_NAME+".add_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message,account,person,action
        person_account=PersonAccount()
        
        person_account.person_id=kwargs['person_id']
        person_account.person_category_id=kwargs['person_category_id']

        result,message,person_account=person_account.save()
        if result==FAILED:
            person_account=None
            return result,message,person_account,None,"ALREADY_EXISTED"

        action="ADDED"
        result=SUCCEED
        message=f"حساب مالی  {person_account.name} با موفقیت به شخص {person_account.person.full_name} اضافه شد."
            
        return result,message,person_account,person_account.person,action

    def remove_account_from_person(self,*args,**kwargs):
        result,message,person_account_id=FAILED,"",0
        if not self.request.user.has_perm(APP_NAME+".add_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message,person_account_id
             
        person_id=kwargs['person_id']
        person_category_id=kwargs['person_category_id']

        person_account=PersonAccount.objects.filter(person_id=person_id).filter(person_category_id=person_category_id).first()
        if person_account is None:
            result=FAILED
            message="حساب مالی وجود ندارد."
            return result,message,person_account_id

        person_account_id=person_account.id
        try:
            person_account.delete()
            person_account_=PersonAccount.objects.filter(id=person_account_id).first()
            if person_account_ is None:
                result=SUCCEED
                message=f"حساب مالی  {person_account.name} با موفقیت از شخص {person_account.person.full_name} حذف شد."
        except:
            message="حذف نشد. "+"ابتدا رویداد های مالی مرتبط را حذف کنید."+"تا تراز فرد صفر شده و هیچ سندی با شخص در ارتباط نباشد."
            return result,message,person_account_id

        return result,message,person_account_id

 
class FinancialYearRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=FinancialYear.objects

        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_person"):
            self.objects=FinancialYear.objects
        elif request.user.is_authenticated:
            self.objects=FinancialYear.objects.filter(profile__user_id=request.user.id)
                 
        else:
            self.objects=FinancialYear.objects.filter(pk=0)

        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
     
    def financial_year(self,*args, **kwargs):
        if "financial_year_id" in kwargs and kwargs["financial_year_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_year_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "date" in kwargs and kwargs["date"] is not None:
            return self.objects.filter(start_date_lte=kwargs['date']).filter(end_date_gte=kwargs['date']).first() 
        
    def current_financial_year(self):
        return self.objects.filter(in_progress=True).first()
   
    def add_financial_year(self,*args,**kwargs):
        result,message,financial_year,financial_years=FAILED,"",None,[]
        if not self.request.user.has_perm(APP_NAME+".add_financialyear"):
            message="دسترسی غیر مجاز"
            return result,message,financial_year,financial_years

        financial_year=FinancialYear() 

 

    
        if 'start_date' in kwargs:
            year=kwargs['start_date'][:2]
            if year=="13" or year=="14":
                kwargs['start_date']=PersianCalendar().to_gregorian(kwargs["start_date"])
            financial_year.start_date=kwargs['start_date']

            
        if 'end_date' in kwargs:
            year=kwargs['end_date'][:2]
            if year=="13" or year=="14":
                kwargs['end_date']=PersianCalendar().to_gregorian(kwargs["end_date"])
            financial_year.end_date=kwargs['end_date']

 

        if 'name' in kwargs:
            financial_year.name=kwargs["name"] 

        if 'description' in kwargs:
            financial_year.description=kwargs["description"] 

        if 'status' in kwargs:
            financial_year.status=kwargs["status"] 
 
        if(len(FinancialYear.objects.filter(name=financial_year.name))>0):
            financial_year=None
            message="نام وارد شده تکراری است."
            result=FAILED
            return result,message,financial_year,financial_years

 
        (result,message,financial_year)=financial_year.save()
        if result==FAILED:
            return result,message,financial_year,financial_years
        
        financial_years=FinancialYear.objects.order_by("start_date")
 
        return result,message,financial_year,financial_years
 
      
class PersonCategoryRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=PersonCategory.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
 
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
     

     
    def delete_all(self,*args,**kwargs):
        
        if not self.request.user.has_perm(APP_NAME+".delete_personcategory"):
            message="دسترسی غیر مجاز"
            return result,message
        PersonCategory.objects.all().delete()
        
        result=SUCCEED
        message="همه حذف شدند."
        return result,message

    def person_category(self,*args, **kwargs):
        if "person_id" in kwargs and kwargs["person_id"] is not None:
            return self.objects.filter(pk=kwargs['person_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
        if "title" in kwargs and kwargs["title"] is not None:
            return self.objects.filter(title=kwargs['title']).first()
             
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
       
    def add_person_category(self,*args,**kwargs):
        result,message,person_category=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_personcategory"):
            message="دسترسی غیر مجاز"
            return result,message,person_category
         
        person_category=PersonCategory()
          
        if 'account_code' in kwargs:
            account_code=kwargs["account_code"]
            person_category.account=Account.objects.filter(code=account_code).first()

          
         
        if 'title' in kwargs:
            if len(PersonCategory.objects.filter(title=kwargs["title"]))>0:
                message="نام وارد شده تکراری می باشد."
                return result,message,None
            person_category.title=kwargs["title"]
        if 'priority' in kwargs:
            person_category.priority=kwargs["priority"] 
        person_category.save()
        result=SUCCEED
        message="دسته بندی جدید برای اشخاص با موفقیت اضافه گردید."
        # (result,message,person_category)=person_category.save()
        # (result,message,person_category)=person_category.save()
        return result,message,person_category

    def initial_default_person_categories(self,*args,**kwargs):
        result=SUCCEED
        message=""
        person_categories,persons=default_persons()
        for person_category in person_categories:
            result,message,new_person_category=self.add_person_category(**person_category)
                
        for person in persons:
            from authentication.repo import PersonRepo
            result,message,new_person=PersonRepo(request=self.request).add_person(**person)

        return result,message       



class ProductSpecificationRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        
        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_event"):
            self.objects=ProductSpecification.objects
        elif request.user.is_authenticated:
            accs=[]
            for person in Person.objects.filter(profile__user_id=request.user.id):

                my_accounts=AccountRepo(request=request).my_accounts
                for acc in my_accounts:
                    accs.append(acc.id)
            self.objects=Event.objects.filter(Q(bedehkar_id__in=accs)|Q(bestankar_id__in=accs))
        else:
            self.objects=Event.objects.filter(pk=0)

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
        return objects.all()
    def product_specification(self,*args, **kwargs):
        if "product_specification_id" in kwargs:
            return self.objects.filter(pk=kwargs['product_specification_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

                     
    def add_product_specification(self,*args, **kwargs):
        product_specification,message,result,deleted_id=(None,"",FAILED,0)
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"productspecification"):
        if not self.request.user.has_perm(APP_NAME+".add_account"):
            message="دسترسی غیر مجاز"
            return result,message,product_specification,deleted_id
        # if len(Account.objects.filter(title=kwargs['title']))>0:
        #     message="از قبل حسابی با همین عنوان ثبت شده است."
        #     return product_specification,message,result
        product_specification=ProductSpecification()
        
        if 'product_id' in kwargs:
            product_specification.product_id=kwargs['product_id']
        if 'name' in kwargs:
            product_specification.name=kwargs['name']
        if 'value' in kwargs:
            product_specification.value=kwargs['value'] 
        deleted=ProductSpecification.objects.filter(product_id=product_specification.product_id).filter(name=product_specification.name).first()
        if deleted is not None:
            deleted_id=deleted.id
            deleted.delete()
         
        product_specification.save()
        result=SUCCEED
        message="ویژگی جدید با موفقیت اضافه گردید."
         
 
        return result,message,product_specification,deleted_id

 


class ProductRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=Product.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for))
        return objects.all()
    
    def add_product_to_category(self,*args, **kwargs):
        
        result,message,category,product_categories=FAILED,"",None,[]
        if not self.request.user.has_perm(APP_NAME+".change_product"):
            message="دسترسی غیر مجاز"
            return result,message,category,product_categories
        

        product=Product.objects.filter(pk=kwargs['product_id']).first()
        category=Category.objects.filter(pk=kwargs['category_id']).first()
        if category is not None and product is not None:
            if product in category.products.all():
                category.products.remove(product.id)
                message="حذف شد"
            else:
                category.products.add(product.id)
                message="اضافه شد"
            result=SUCCEED
            product_categories=product.category_set.all()
        return result,message,category,product_categories
       
    def product(self,*args, **kwargs):
        if "product_id" in kwargs and kwargs["product_id"] is not None:
            return self.objects.filter(pk=kwargs['product_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(barcode=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
           
    def add_product(self,*args,**kwargs):
        result,message,product=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_product"):
            message="دسترسی غیر مجاز"
            return result,message,product
        if len(Product.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای کالای جدید"
            return result,message,product

        product=Product() 

        if 'title' in kwargs:
            product.title=kwargs["title"]

            
        if 'brand_id' in kwargs:
            product.brand_id=kwargs["brand_id"]

            
        if 'model' in kwargs:
            product.model=kwargs["model"]
         
        if 'barcode' in kwargs and kwargs["barcode"] is not None and not kwargs["barcode"]=="":
            product.barcode=kwargs["barcode"]
        
        if product.barcode is not None and len(product.barcode)>0:
            
            if len(Product.objects.filter(barcode=product.barcode))>0:
                message="بارکد تکراری برای کالای جدید"
                return result,message,None

        (result,message,product)=product.save()
        if 'unit_price' in kwargs and kwargs['unit_price'] is not None:
            if 'unit_name' in kwargs and kwargs['unit_name'] is not None:
                if 'coef' in kwargs and kwargs['coef'] is not None:
                    ili_unit=InvoiceLineItemUnit()
                    ili_unit.unit_name=kwargs["unit_name"]
                    ili_unit.coef=kwargs["coef"]
                    ili_unit.unit_price=kwargs["unit_price"]
                    ili_unit.invoice_line_item_id=product.id
                    ili_unit.default=True
                    ili_unit.save()

                 

        if 'category_id' in kwargs:
            category_id=kwargs["category_id"]
            category=Category.objects.filter(pk=category_id).first()
            if category is not None:
                category.products.add(product.id)
        return result,message,product
 

    def import_products_from_excel(self,*args,**kwargs):
        result,message,products=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # products=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        count=kwargs['count']
        products_to_import=[]
        START_ROW=EXCEL_PRODUCTS_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            product={}
            product['id']=ws['A'+str(i)].value
            product['title']=ws['B'+str(i)].value
            product['barcode']=ws['C'+str(i)].value
            product['unit_name']=ws['D'+str(i)].value
            product['unit_price']=ws['E'+str(i)].value
            product['thumbnail_origin']=ws['F'+str(i)].value
            # product['thumbnail_origin']=ws['F'+str(i)].value
            if product['title'] is not None and not product['title']=="":
                products_to_import.append(product) 
        modified=added=0
        for product in products_to_import:
            old_product=Product.objects.filter(title=product["title"]).filter(barcode=product["barcode"]).first()
            if old_product is not None:
                old_product.title=product["title"]
                # old_product.unit_name=product["unit_name"]
                old_product.thumbnail_origin=product["thumbnail_origin"]
                # old_product.unit_price=product["unit_price"] 
                # old_product.thumbnail_origin=product["thumbnail_origin"] 
                old_product.save()
                modified+=1
            else:
                try:
                    result,message,new_product=self.add_product(title=product["title"],barcode=product["barcode"],unit_name=product["unit_name"],unit_price=product["unit_price"] ,coef=1)
                    products.append(new_product)
                except:
                    pass
                # new_product.title=product["title"]
                # new_product.barcode=product["barcode"]
                # new_product.unit_name=product["unit_name"]
                # new_product.unit_price=product["unit_price"] 
                # new_product.save()
                if result==SUCCEED:
                    added+=1
        result=SUCCEED
        message=f"""{added} محصول اضافه شد.
                    <br>
                    {modified} محصول ویرایش شد. """
        products=self.list()
        return result,message,products


class InvoiceLineItemRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=InvoiceLineItem.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for))
        return objects.all()
    
         
    def invoice_line_item(self,*args, **kwargs):
        if "invoice_line_item_id" in kwargs and kwargs["invoice_line_item_id"] is not None:
            return self.objects.filter(pk=kwargs['invoice_line_item_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(barcode=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
           
    def add_invoice_line_item(self,*args,**kwargs):
        result,message,invoice_line_item=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_invoice_line_item"):
            message="دسترسی غیر مجاز"
            return result,message,invoice_line_item
        if len(InvoiceLineItem.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای مورد جدید"
            return result,message,invoice_line_item

        invoice_line_item=InvoiceLineItem() 

        if 'title' in kwargs:
            invoice_line_item.title=kwargs["title"]
        if 'unit_price' in kwargs:
            invoice_line_item.unit_price=kwargs["unit_price"]
        if 'unit_price' in kwargs:
            invoice_line_item.unit_price=kwargs["unit_price"]
            
        if 'barcode' in kwargs and kwargs["barcode"] is not None and not kwargs["barcode"]=="":
            invoice_line_item.barcode=kwargs["barcode"]
        
        if 'unit_name' in kwargs:
            invoice_line_item.unit_name=kwargs["unit_name"]

            
        if invoice_line_item.barcode is not None and len(invoice_line_item.barcode)>0:
            
            if len(Product.objects.filter(barcode=invoice_line_item.barcode))>0:
                message="بارکد تکراری برای کالای جدید"
                return result,message,None

        (result,message,invoice_line_item)=invoice_line_item.save()
        if 'category_id' in kwargs:
            category_id=kwargs["category_id"]
            category=Category.objects.filter(pk=category_id).first()
            if category is not None:
                category.invoice_line_items.add(invoice_line_item.id)
        coef=1
        if 'coef' in kwargs:
            coef=kwargs["coef"]
        if invoice_line_item.unit_price>0:
            InvoiceLineItemUnitRepo(request=self.request).add_invoice_line_item_unit(invoice_line_item_id=invoice_line_item.id,unit_price=invoice_line_item.unit_price,unit_name=invoice_line_item.unit_name,coef=coef)
        return result,message,invoice_line_item
 
 

class BankAccountRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=BankAccount.objects
       

        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "bank_id" in kwargs:
            bank_id=kwargs["bank_id"]
            objects=objects.filter(Q(bank_id=bank_id))
  
        if "person_id" in kwargs:
            person_id=kwargs["person_id"]
            objects=objects.filter(Q(person_id=person_id))
       
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(full_name__contains=search_for) | Q(melli_code__contains=search_for) | Q(code=search_for))
        return objects.all()
     
    def bank_account(self,*args, **kwargs):
        if "bank_account_id" in kwargs and kwargs["bank_account_id"] is not None:
            return self.objects.filter(pk=kwargs['bank_account_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
             
    def add_bank_account(self,*args,**kwargs):
        result,message,bank_account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_bankaccount"):
            message="دسترسی غیر مجاز"
            return result,message,bank_account

        bank_account=BankAccount() 

 

        if BankAccount.objects.filter(title=kwargs['title']).first() is not None:
            message="نام وارد شده تکراری است."
            bank_account=None
            return result,message,bank_account

        if BankAccount.objects.filter(code=kwargs['code']).first() is not None:
            message="کد وارد شده تکراری است."
            bank_account=None
            return result,message,bank_account

  
        if 'title' in kwargs:
            bank_account.title=kwargs["title"] 
            bank_account.name=kwargs["title"] 


            
  
        if 'card_no' in kwargs:
            bank_account.card_no=kwargs["card_no"] 

            
  
        if 'shaba_no' in kwargs:
            bank_account.shaba_no=kwargs["shaba_no"] 

            
  
        if 'account_no' in kwargs:
            bank_account.account_no=kwargs["account_no"] 

        if 'bank_id' in kwargs:
            bank_account.bank_id=kwargs["bank_id"] 

        if 'person_id' in kwargs:
            bank_account.person_id=kwargs["person_id"]

        if 'code' in kwargs:
            bank_account.code=kwargs["code"] 

        if 'parent_code' in kwargs:
            parent=Account.objects.filter(code=kwargs['parent_code']).first()
            bank_account.parent=parent
        
        bank_account.save() 
        if bank_account.id is not None:
            message="حساب بانکی جدید با موفقیت اضافه شد."
            result=SUCCEED
        return result,message,bank_account

 
class FinancialDocumentRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=FinancialDocument.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for))
        return objects.all()
    
       
    def financial_document(self,*args, **kwargs):
        if "financial_document_id" in kwargs and kwargs["financial_document_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_document_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(barcode=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
           
    def add_financial_document(self,*args,**kwargs):
        result,message,financial_document=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_financial_document"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document
        if len(FinancialDocument.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای کالای جدید"
            return result,message,financial_document

        financial_document=FinancialDocument() 

        if 'title' in kwargs:
            financial_document.title=kwargs["title"]
         
        if 'barcode' in kwargs and kwargs["barcode"] is not None and not kwargs["barcode"]=="":
            financial_document.barcode=kwargs["barcode"]
        
        if financial_document.barcode is not None and len(financial_document.barcode)>0:
            
            if len(FinancialDocument.objects.filter(barcode=financial_document.barcode))>0:
                message="بارکد تکراری برای کالای جدید"
                return result,message,None

        (result,message,financial_document)=financial_document.save()
        if 'unit_price' in kwargs:
            if 'unit_name' in kwargs:
                if 'coef' in kwargs:
                    ili_unit=InvoiceLineItemUnit()
                    ili_unit.unit_name=kwargs["unit_name"]
                    ili_unit.coef=kwargs["coef"]
                    ili_unit.unit_price=kwargs["unit_price"]
                    ili_unit.invoice_line_item_id=financial_document.id
                    ili_unit.default=True
                    ili_unit.save()

                 

        if 'category_id' in kwargs:
            pass
            # category_id=kwargs["category_id"]
            # category=Category.objects.filter(pk=category_id).first()
            # if category is not None:
            #     category.financial_documents.add(financial_document.id)
        coef=1 
        return result,message,financial_document
 

    def import_financial_documents_from_excel(self,*args,**kwargs):
        result,message,financial_documents=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # financial_documents=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        count=kwargs['count']
        financial_documents_to_import=[]

        for i in range(2,count+2):
            financial_document={}
            financial_document['id']=ws['A'+str(i)].value
            financial_document['title']=ws['B'+str(i)].value
            financial_document['code']=ws['C'+str(i)].value
            financial_document['unit_name']=ws['D'+str(i)].value
            financial_document['unit_price']=ws['E'+str(i)].value
            financial_document['thumbnail_origin']=ws['F'+str(i)].value
            # financial_document['thumbnail_origin']=ws['F'+str(i)].value
            if financial_document['title'] is not None and not financial_document['title']=="":
                financial_documents_to_import.append(financial_document) 
        modified=added=0
        for financial_document in financial_documents_to_import:
            old_financial_document=FinancialDocument.objects.filter(title=financial_document["title"]).filter(code=financial_document["code"]).first()
            if old_financial_document is not None:
                old_financial_document.title=financial_document["title"]
                old_financial_document.unit_name=financial_document["unit_name"]
                old_financial_document.thumbnail_origin=financial_document["thumbnail_origin"]
                old_financial_document.unit_price=financial_document["unit_price"] 
                # old_financial_document.thumbnail_origin=financial_document["thumbnail_origin"] 
                old_financial_document.save()
                modified+=1
            else:
                new_financial_document=FinancialDocument()
                new_financial_document.title=financial_document["title"]
                new_financial_document.barcode=financial_document["code"]
                new_financial_document.unit_name=financial_document["unit_name"]
                new_financial_document.unit_price=financial_document["unit_price"] 
                new_financial_document.save()
                added+=1
        result=SUCCEED
        message=f"""{added} محصول اضافه شد.
                    <br>
                    {modified} محصول ویرایش شد. """

        return result,message,financial_documents



class BrandRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=Brand.objects
       

        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(full_name__contains=search_for) | Q(melli_code__contains=search_for) | Q(code=search_for))
        return objects.all()
     
    def brand(self,*args, **kwargs):
        if "brand_id" in kwargs and kwargs["brand_id"] is not None:
            return self.objects.filter(pk=kwargs['brand_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
    
    def initial_default_brands(self,*args, **kwargs):
      
        brands_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_brand"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for brand in default_brands():
            new_brand=Brand(name=brand["name"])
            # new_account=Account(parent=parent_account,**kwargs)
            new_brand.save()
            brands_counter+=1

        message=f"{brands_counter} بانک با موفقیت اضافه شد."
        return result,message

        
    def delete_all(self,*args, **kwargs):
      
        brands_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".delete_brand"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for brand in Brand.objects.all():
            brand.delete()
            brands_counter+=1

        message=f"<p>{brands_counter} بانک با موفقیت حذف شد.</p>"
        return result,message
    
    
    def add_brand(self,*args,**kwargs):
        result,message,brand=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_brand"):
            message="دسترسی غیر مجاز"
            return result,message,brand

        brand=Brand() 

 

        if Brand.objects.filter(name=kwargs['name']).first() is not None:
            message="نام وارد شده تکراری است."
            brand=None
            return result,message,brand

  
        if 'name' in kwargs:
            brand.name=kwargs["name"] 
        brand.save()       
        message="بانک جدید با موفقیت اضافه شد."
        result=SUCCEED
        return result,message,brand

 


class BankRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=Bank.objects
       

        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(full_name__contains=search_for) | Q(melli_code__contains=search_for) | Q(code=search_for))
        return objects.all()
     
    def bank(self,*args, **kwargs):
        if "bank_id" in kwargs and kwargs["bank_id"] is not None:
            return self.objects.filter(pk=kwargs['bank_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
    
    def initial_default_banks(self,*args, **kwargs):
      
        banks_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_bank"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for bank in default_banks():
            new_bank=Bank(name=bank["name"])
            # new_account=Account(parent=parent_account,**kwargs)
            new_bank.save()
            banks_counter+=1

        message=f"{banks_counter} بانک با موفقیت اضافه شد."
        return result,message

        
    def delete_all(self,*args, **kwargs):
      
        banks_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".delete_bank"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for bank in Bank.objects.all():
            bank.delete()
            banks_counter+=1

        message=f"<p>{banks_counter} بانک با موفقیت حذف شد.</p>"
        return result,message
    
    
    def add_bank(self,*args,**kwargs):
        result,message,bank=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_bank"):
            message="دسترسی غیر مجاز"
            return result,message,bank

        bank=Bank() 

 

        if Bank.objects.filter(name=kwargs['name']).first() is not None:
            message="نام وارد شده تکراری است."
            bank=None
            return result,message,bank

  
        if 'name' in kwargs:
            bank.name=kwargs["name"] 
        bank.save()       
        message="بانک جدید با موفقیت اضافه شد."
        result=SUCCEED
        return result,message,bank

 
class ServiceRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=Service.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for))
        return objects.all()
    
    def add_product_to_category(self,*args, **kwargs):
        
        result,message,category,product_categories=FAILED,"",None,[]
        if not self.request.user.has_perm(APP_NAME+".change_product"):
            message="دسترسی غیر مجاز"
            return result,message,category,product_categories
        

        product=Product.objects.filter(pk=kwargs['product_id']).first()
        category=Category.objects.filter(pk=kwargs['category_id']).first()
        if category is not None and product is not None:
            if product in category.products.all():
                category.products.remove(product.id)
                message="حذف شد"
            else:
                category.products.add(product.id)
                message="اضافه شد"
            result=SUCCEED
            product_categories=product.category_set.all()
        return result,message,category,product_categories
       
    def service(self,*args, **kwargs):
        if "service_id" in kwargs and kwargs["service_id"] is not None:
            return self.objects.filter(pk=kwargs['service_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(barcode=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
            
    def add_service(self,*args,**kwargs):
        result,message,service=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_service"):
            message="دسترسی غیر مجاز"
            return result,message,service
        if len(Service.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای سرویس جدید"
            return result,message,service

        service=Service() 

        if 'title' in kwargs:
            service.title=kwargs["title"]
         
    

        (result,message,service)=service.save()
        if 'unit_price' in kwargs:
            if 'unit_name' in kwargs:
                if 'coef' in kwargs:
                    ili_unit=InvoiceLineItemUnit()
                    ili_unit.unit_name=kwargs["unit_name"]
                    ili_unit.coef=kwargs["coef"]
                    ili_unit.unit_price=kwargs["unit_price"]
                    ili_unit.invoice_line_item_id=service.id
                    ili_unit.default=True
                    ili_unit.save()

                 

        if 'category_id' in kwargs:
            pass
            # category_id=kwargs["category_id"]
            # category=Category.objects.filter(pk=category_id).first()
            # if category is not None:
            #     category.services.add(service.id)
        coef=1 
        return result,message,service
 
    def import_services_from_excel(self,*args,**kwargs):
        result,message,services=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # services=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        count=kwargs['count']
        services_to_import=[]

        START_ROW=EXCEL_SERVICES_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            service={}
            service['id']=ws['A'+str(i)].value
            service['title']=ws['B'+str(i)].value
            service['unit_name']=ws['D'+str(i)].value
            service['unit_price']=ws['E'+str(i)].value
            service['thumbnail_origin']=ws['F'+str(i)].value
            # service['thumbnail_origin']=ws['F'+str(i)].value
            if service['title'] is not None and not service['title']=="":
                services_to_import.append(service) 
        modified=added=0
        for service in services_to_import:
            old_service=Service.objects.filter(title=service["title"]).first()
            if old_service is not None:
                old_service.title=service["title"]
                # old_service.unit_name=service["unit_name"]
                old_service.thumbnail_origin=service["thumbnail_origin"]
                # old_service.unit_price=service["unit_price"] 
                # old_service.thumbnail_origin=service["thumbnail_origin"] 
                old_service.save()
                modified+=1
            else:
                result,message,new_service=self.add_service(title=service["title"],unit_name=service["unit_name"],unit_price=service["unit_price"] ,coef=1)
                # new_service.title=service["title"]
                # new_service.barcode=service["barcode"]
                # new_service.unit_name=service["unit_name"]
                # new_service.unit_price=service["unit_price"] 
                # new_service.save()
                if result==SUCCEED:
                    added+=1
        result=SUCCEED
        message=f"""{added} سرویس اضافه شد.
                    <br>
                    {modified} سرویس ویرایش شد. """
        services=self.list()
        return result,message,services


class FinancialDocumentRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=FinancialDocument.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    
    def list(self,*args, **kwargs):
        objects=self.objects
        if "financial_year_id" in kwargs:
            objects=objects.filter(financial_year_id=kwargs['financial_year_id']) 
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
        return objects.all()
    
    def financial_document(self,*args, **kwargs):
        if "financial_document_id" in kwargs:
            return self.objects.filter(pk=kwargs['financial_document_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

            
    def add_financial_document(self,*args, **kwargs):
        result,message,financial_document=FAILED,"",None
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"financialdocument"):
        if not self.request.user.has_perm(APP_NAME+".add_financialdocument"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document
        
        f_year=FinancialYear.objects.filter(status=FinancialYearStatusEnum.IN_PROGRESS).first()
        if f_year is None:
            url=reverse(APP_NAME+":financial_years")
            message="سال مالی فعال وجود ندارد. ابتدا ایجاد کنید."+"<br>"+"<a href='"+url+"'>سال های مالی</a>"
            return result,message,financial_document

        financial_document=FinancialDocument(financial_year_id=f_year.id)

        if 'title' in kwargs:
            financial_document.title=kwargs['title']
     
        
        # if 'financial_year_id' in kwargs:
        #     payment.financial_year_id=kwargs['financial_year_id']
        # else:
        #     payment.financial_year_id=FinancialYear.get_by_date(date=payment.transaction_datetime).id
        result,message,financial_document=financial_document.save()
        if result==SUCCEED:
            message="با موفقیت اضافه شد."
            me_profile=ProfileRepo(request=self.request).me
            new_log={}
            new_log['title']="سند مالی جدید "+" : "+financial_document.title
            new_log['app_name']=APP_NAME
            new_log['url']=financial_document.get_absolute_url()
            new_log['profile']=me_profile
            new_log['description']="سند مالی جدید با موفقیت اضافه گردید."
            LogRepo(request=self.request).add_log(**new_log)
        return result,message,financial_document


class FinancialDocumentLineRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=ProfileRepo(request=request).me
        self.objects=FinancialDocumentLine.objects
        
    def list(self,*args, **kwargs):
        objects=self.objects
        if "start_date" in kwargs and kwargs["start_date"] is not None :
            start_date=kwargs["start_date"]
            objects=objects.filter(date_time__gte=start_date) 
        if "end_date" in kwargs and kwargs["end_date"] is not None :
            end_date=kwargs["end_date"]
            objects=objects.filter(date_time__lte=end_date) 
        if "search_for" in kwargs and kwargs["search_for"] is not None and len(kwargs["search_for"])>0 :
            objects=objects.filter(Q(title__contains=kwargs['search_for'])|Q(event__title__contains=kwargs['search_for']) )
        if "amount" in kwargs and kwargs["amount"] is not None and kwargs["amount"]>0 :
            objects=objects.filter(Q(bedehkar=kwargs['amount']) | Q(bestankar=kwargs['amount']) )

            
        if "bestankar" in kwargs and kwargs["bestankar"] is not None:
            objects=objects.filter(Q(bestankar=kwargs['bestankar']) )

        if "bedehkar" in kwargs and kwargs["bedehkar"] is not None:
            objects=objects.filter(Q(bedehkar=kwargs['bedehkar']) )

        if "account_code" in kwargs and kwargs["account_code"] is not None :
            account_code=kwargs["account_code"]
            account=AccountRepo(request=self.request).account(code=account_code)
            if account is not None:
                objects=objects.filter(account_id=account.id)
        if "account_id" in kwargs and kwargs["account_id"] is not None and kwargs["account_id"]>0 :
            account_id=kwargs["account_id"]
            objects=objects.filter(account_id=account_id)
        if "event_id" in kwargs and kwargs["event_id"] is not None and kwargs["event_id"]>0 :
            event_id=kwargs["event_id"]
            objects=objects.filter(financial_event_id=event_id)

        if "financial_event_id" in kwargs and kwargs["financial_event_id"] is not None and kwargs["financial_event_id"]>0 :
            financial_event_id=kwargs["financial_event_id"]
            objects=objects.filter(financial_event_id=financial_event_id)
        
        if "financial_document_id" in kwargs and kwargs["financial_document_id"] is not None and kwargs["financial_document_id"]>0 :
            financial_document_id=kwargs["financial_document_id"]
            objects=objects.filter(financial_document_id=financial_document_id)
        
        return objects.all().order_by('date_time')

    def financial_document_line(self,*args, **kwargs):
        if "financial_document_line_id" in kwargs and kwargs["financial_document_line_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_document_line_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
    def add_financial_document_line(self,*args, **kwargs):
        financial_document_line,message,result=(None,"",FAILED)
    
        bestankar=kwargs['bestankar']
        bedehkar=kwargs['bedehkar']
        if bedehkar==0 and bestankar==0:
            message="مبلغ بدهکار و بستانکار صفر وارد شده است."
            return result,message,financial_document_line

            
        if bedehkar>0 and bestankar>0:
            message="مبلغ بدهکار و بستانکار ، هر دو وارد شده است."
            return result,message,financial_document_line

        if bedehkar<0 or bestankar<0:
            message="مبلغ بدهکار یا بستانکار منفی وارد شده است."
            return result,message,financial_document_line

        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"accountingdocumentline"):
        if not self.request.user.has_perm(APP_NAME+".add_accountingdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document_line
        
        financial_document_line=FinancialDocumentLine()
        
        if 'title' in kwargs:
            financial_document_line.title=kwargs['title']
        if 'financial_event_id' in kwargs:
            financial_document_line.financial_event_id=kwargs['financial_event_id']
        if 'financial_document_id' in kwargs:
            financial_document_id=kwargs['financial_document_id']
            if int(financial_document_id)==0 and 'financial_document_title' in kwargs:
                result,message,financial_document=FinancialDocumentRepo(request=self.request).add_financial_document(title=kwargs['financial_document_title'])
                financial_document_id=financial_document.id
            financial_document_line.financial_document_id=financial_document_id
        if 'description' in kwargs:
            financial_document_line.description=kwargs['description']
        if 'persian_date_time' in kwargs and kwargs['persian_date_time'] is not None and not kwargs['persian_date_time']=='':
            persian_date_time=kwargs['persian_date_time']
            date_time=PersianCalendar().to_gregorian(persian_date_time)
            # date_time=date_time,persian_date_time=kwargs['persian_date_time'])
            # financial_document_line.date_time=date_time
        if 'bestankar' in kwargs  :
            financial_document_line.bestankar=kwargs['bestankar']
        if 'bedehkar' in kwargs :
            financial_document_line.bedehkar=kwargs['bedehkar'] 
        if 'date_time' in kwargs :

            date_time=kwargs['date_time']
            year=date_time[:2]
            if year=="13" or year=="14":
                date_time=PersianCalendar().to_gregorian(kwargs["date_time"])
            financial_document_line.date_time=date_time 

        if 'account_code' in kwargs and kwargs['account_code'] is not None:
            account=AccountRepo(request=self.request).account(code=kwargs['account_code']) 
            if account is not None:
                financial_document_line.account=account
        if 'account_id' in kwargs and kwargs['account_id'] is not None:
            financial_document_line.account_id=kwargs['account_id'] 
        
        
        # if 'financial_year_id' in kwargs:
        #     payment.financial_year_id=kwargs['financial_year_id']
        # else:
        #     payment.financial_year_id=FinancialYear.get_by_date(date=payment.transaction_datetime).id

        if financial_document_line.account.nature==AccountNatureEnum.ONLY_BESTANKAR and financial_document_line.bedehkar>0:
            message=financial_document_line.account.name+" ماهیت فقط بستانکار دارد"
            financial_document_line=None
            return result,message,financial_document_line
        if financial_document_line.account.nature==AccountNatureEnum.ONLY_BEDEHKAR and financial_document_line.bestankar>0:
            message=financial_document_line.account.name+" ماهیت فقط بدهکار دارد"
            financial_document_line=None
            return result,message,financial_document_line


        result,message,financial_document_line=financial_document_line.save()
        if result==FAILED:
            return result,message,financial_document_line
        # financial_document_line.account.normalize_total()
        result=SUCCEED
        message="با موفقیت اضافه گردید."
         

        me_profile=ProfileRepo(request=self.request).me
        new_log={}
        new_log['title']="خط سند مالی جدید "
        new_log['app_name']=APP_NAME
        new_log['url']=financial_document_line.get_absolute_url()
        new_log['profile']=me_profile
        new_log['description']="خط سند مالی جدید با موفقیت اضافه گردید."
        LogRepo(request=self.request).add_log(**new_log)
        return result,message,financial_document_line
    
    def delete_all(self,*args,**kwargs):
        
        if not self.request.user.has_perm(APP_NAME+".delete_accountingdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message
        FinancialDocumentLine.objects.all().delete() 
                   
        result=SUCCEED
        message="همه حذف شدند."
        return result,message
    
    
    
    def add_event_financial_document_line(self,*args, **kwargs):
        result,message,financial_document_line=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_financialdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message,accounting_document_line

        event_id=kwargs['event_id']
        bestankar=kwargs['bestankar']
        bedehkar=kwargs['bedehkar']
        account_code=kwargs['account_code']
        financial_document_id=kwargs['financial_document_id']
        financial_document_title=kwargs['financial_document_title']

        account_repo=AccountRepo(request=self.request)
        event_repo=FinancialEventRepo(request=self.request)
        financial_document_repo=FinancialDocumentRepo(request=self.request)

        event=event_repo.event(pk=event_id)
        account=account_repo.account(code=account_code)

        if financial_document_id==0:
            result,message,financial_document=financial_document_repo.add_financial_document(title=financial_document_title)
        else:
            financial_document=financial_document_repo.financial_document(pk=financial_document_id)
        if account is not None and financial_document is not None and event is not None:
            financial_document_line=FinancialDocumentLine()
            financial_document_line.event=event
            financial_document_line.bestankar=bestankar
            financial_document_line.bedehkar=bedehkar
            financial_document_line.account=account
            financial_document_line.date_time=event.event_datetime
            financial_document_line.title=event.title
            financial_document_line.financial_document=financial_document
            financial_document_line.save()
            result=SUCCEED
            message="اضافه شد."
            return result,message,financial_document_line
  

class InvoiceRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Invoice.objects.filter(id=0)
        profile=ProfileRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_invoice"):
                self.objects=Invoice.objects
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            codeee=str(filter_number(search_for))
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(code=codeee) )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.all()
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def invoice(self,*args, **kwargs):
        if "invoice_id" in kwargs and kwargs["invoice_id"] is not None:
            return self.objects.filter(pk=kwargs['invoice_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
         
       

    def add_invoice(self,*args,**kwargs):
        result,message,invoice=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_invoice"):
            message="دسترسی غیر مجاز"
            return result,message,invoice

        invoice=Invoice()
        if 'title' in kwargs:
            invoice.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                invoice.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            invoice.color=kwargs["color"]
        if 'code' in kwargs:
            invoice.code=kwargs["code"]
        if 'priority' in kwargs:
            invoice.priority=kwargs["priority"]
        if 'bedehkar_id' in kwargs:
            invoice.bedehkar_id=kwargs["bedehkar_id"]
        if 'bestankar_id' in kwargs:
            invoice.bestankar_id=kwargs["bestankar_id"]
        if 'event_datetime' in kwargs:
            
            year=kwargs['event_datetime'][:2]
            if year=="13" or year=="14":
                kwargs['event_datetime']=PersianCalendar().to_gregorian(kwargs["event_datetime"])
            invoice.event_datetime=kwargs["event_datetime"]

        if 'type' in kwargs:
            invoice.type=kwargs["type"]

           
        (result,message,invoice)=invoice.save()
        return result,message,invoice


class FinancialEventRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=FinancialEvent.objects.filter(id=0)
        profile=ProfileRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_invoice"):
                self.objects=FinancialEvent.objects
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            codeee=str(filter_number(search_for))
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(code=codeee) )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.all()
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def financial_event(self,*args, **kwargs):
        if "financial_event_id" in kwargs and kwargs["financial_event_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_event_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "event_id" in kwargs and kwargs["event_id"] is not None:
            return self.objects.filter(pk=kwargs['event_id']).first() 
         
       
    def delete_all(self,*args, **kwargs):
        result,message=FAILED,""
        if not self.request.user.has_perm(APP_NAME+".delete_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message
        financial_events=FinancialEvent.objects.all()
        financial_events.delete()
        result=SUCCEED
        message='همه رویداد ها حذف شدند.'
        return result,message
    def add_financial_event(self,*args,**kwargs):
        result,message,event=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message,event

        financial_event=FinancialEvent()
        if 'bedehkar_id' in kwargs:
            financial_event.bedehkar_id=kwargs["bedehkar_id"]
        if 'bestankar_id' in kwargs:
            financial_event.bestankar_id=kwargs["bestankar_id"]
        if 'title' in kwargs:
            financial_event.title=kwargs["title"]
        if 'description' in kwargs:
            financial_event.description=kwargs["description"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                financial_event.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            financial_event.color=kwargs["color"]
        if 'amount' in kwargs:
            financial_event.amount=kwargs["amount"]
        if 'priority' in kwargs:
            financial_event.priority=kwargs["priority"]
        if 'type' in kwargs:
            financial_event.type=kwargs["type"]
        if 'event_datetime' in kwargs:
            event_datetime=kwargs["event_datetime"]
            financial_event.event_datetime=event_datetime

           
            year=event_datetime[:2]
            if year=="13" or year=="14":
                event_datetime=PersianCalendar().to_gregorian(event_datetime)
            financial_event.event_datetime=event_datetime

        (result,message,financial_event)=financial_event.save()
        return result,message,financial_event
 
class CategoryRepo():

    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_categorys=[]
        self.request=request
        self.objects=Category.objects.filter(id=0)
        profile=ProfileRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_category"):
                self.objects=Category.objects
                self.my_categorys=self.objects
                
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(title__contains=search_for))
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.all()
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def category(self,*args, **kwargs):
        if "category_id" in kwargs and kwargs["category_id"] is not None:
            return self.objects.filter(pk=kwargs['category_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            code=kwargs['code']
            code=kwargs['code']

            category= self.objects.filter(code=code).first()
            return category
        if "category_code" in kwargs and kwargs["category_code"] is not None:
            a= self.objects.filter(code=kwargs['category_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['category_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
          
    def set_priority(self,*args, **kwargs):
        result,message,priority=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".change_category"):
            return result,message,category_tags
        priority=kwargs['priority']
        category_id=kwargs['category_id']
        category=Category.objects.filter(pk=category_id).first()
        if category is not None:
            category.priority=priority
            category.save()
        result=SUCCEED
        return result,message,priority

    def set_category_parent(self,*args, **kwargs):
        result,message,category,parent=FAILED,"",None,None
        category=self.category(*args,**kwargs)
        parent=None
        parent_code=kwargs["parent_code"]
        parent=self.category(category_code=parent_code)
        category.parent=parent
        category.save()
        result=SUCCEED
        message="با موفقیت تغییر یافت"
        return result,message,category,parent

    def initial_default_categories(self,*args, **kwargs):
        category_group_counter=0
        basic_categorys_counter=0
        moein_categorys_counter=0
        moein2_categorys_counter=0
        tafsili_categorys_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_category"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        
        categorys=default_categorys() 

        for category in categorys:
            parent_category=None
            if 'parent_code' in category:
                parent_category=Category.objects.filter(code=category["parent_code"]).first()
            new_category=Category(name=category["name"],color=category["color"],code=category['code'],priority=category['priority'],parent=parent_category)
            # new_category=Category(parent=parent_category,**kwargs)
            new_category.save()
            # category_group_counter+=1
            # basic_categorys_counter+=1   
 
                                  
        categorys=Category.objects
        category_group_counter=len(categorys.filter(type=CategoryTypeEnum.GROUP))
        basic_categorys_counter=len(categorys.filter(type=CategoryTypeEnum.BASIC))
        moein_categorys_counter=len(categorys.filter(type=CategoryTypeEnum.MOEIN_1))
        moein2_categorys_counter=len(categorys.filter(type=CategoryTypeEnum.MOEIN_2))
        tafsili_categorys_counter=len(categorys.filter(Q(type=CategoryTypeEnum.TAFSILI_1)|Q(type=CategoryTypeEnum.TAFSILI_2)|Q(type=CategoryTypeEnum.TAFSILI_3)|Q(type=CategoryTypeEnum.TAFSILI_4)))
                                                        

        if result==SUCCEED:
            message="با موفقیت اضافه گردید."
        message+=f"<br>{category_group_counter}   گروه حساب" 
        message+=f"<br>{basic_categorys_counter}   حساب  کل " 
        message+=f"<br>{moein_categorys_counter}  حساب معین سطح یک " 
        message+=f"<br>{moein2_categorys_counter}  حساب معین سطح دو " 
        message+=f"<br>{tafsili_categorys_counter}  حساب تفصیلی " 

        me_profile=ProfileRepo(request=self.request).me
        new_log={}
        new_log['title']="افزودن حساب های پیش فرض"
        new_log['app_name']=APP_NAME
        new_log['profile']=me_profile
        new_log['description']="حساب های پیش فرض اضافه شدند."
        LogRepo(request=self.request).add_log(**new_log)
        return result,message
 
    def add_category(self,*args,**kwargs):
        result,message,category=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_category"):
            message="دسترسی غیر مجاز"
            return result,message,category
        category=Category()
        if 'title' in kwargs:
            category.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                category.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            category.color=kwargs["color"]
        if 'priority' in kwargs and kwargs['priority'] is not None:
            category.priority=kwargs["priority"]
        
        (result,message,category)=category.save()
        return result,message,category
    def add_product_to_category(self,*args, **kwargs):
        result,message,product_categories=FAILED,'',[]
            
        if not self.request.user.has_perm(APP_NAME+".add_category"):
            message="دسترسی غیر مجاز"
            return result,message,product_categories
        # product_id=0
        # category_id=0
        # if 'category_id' in kwargs:
        #     category_id=kwargs["category_id"]
        # if 'product_id' in kwargs:
        #     product_id=kwargs["product_id"]
        product=ProductRepo(request=self.request).product(*args, **kwargs)
        category=self.category(*args, **kwargs)
        if product is None:
            message="کالایی پیدا نشد"
            return result,message,product_categories
        if category is None:
            message="دسته بندی پیدا نشد"
            return result,message,product_categories
        if product in category.products.all():
            message='با موفقیت کالا از این دسته بندی حذف شد.'
            result=SUCCEED
            category.products.remove(product.id)
            product_categories=product.category_set.all()
            return result,message,product_categories
        category.products.add(product.id)
        result=SUCCEED
        message='با موفقیت کالا به دسته بندی اضافه شد.'
        product_categories=product.category_set.all()
        return result,message,product_categories
    
