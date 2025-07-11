from django.test import TestCase

# Create your tests here.
 def import_products_from_excel(self,*args,**kwargs):
        result,message,products=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # products=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        count=kwargs['count']
        products_to_import=[]

        for i in range(2,count+2):
            product={}
            product['id']=ws['A'+str(i)].value
            product['title']=ws['B'+str(i)].value
            product['code']=ws['C'+str(i)].value
            product['unit_name']=ws['D'+str(i)].value
            product['unit_price']=ws['E'+str(i)].value
            product['thumbnail_origin']=ws['F'+str(i)].value
            # product['thumbnail_origin']=ws['F'+str(i)].value
            if product['title'] is not None and not product['title']=="":
                products_to_import.append(product) 
        modified=added=0
        for product in products_to_import:
            old_product=Product.objects.filter(title=product["title"]).filter(code=product["code"]).first()
            if old_product is not None:
                old_product.title=product["title"]
                old_product.unit_name=product["unit_name"]
                old_product.thumbnail_origin=product["thumbnail_origin"]
                old_product.unit_price=product["unit_price"] 
                # old_product.thumbnail_origin=product["thumbnail_origin"] 
                old_product.save()
                modified+=1
            else:
                new_product=Product()
                new_product.title=product["title"]
                new_product.barcode=product["code"]
                new_product.unit_name=product["unit_name"]
                new_product.unit_price=product["unit_price"] 
                new_product.save()
                added+=1
        result=SUCCEED
        message=f"""{added} محصول اضافه شد.
                    <br>
                    {modified} محصول ویرایش شد. """

        return result,message,products

